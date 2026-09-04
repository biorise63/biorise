#!/usr/bin/env python3
"""
Конвертирует уже собранный yandex-doctors-feed/biorise-doctors-feed.xlsx
в YML-фид врачей (реальная схема из официального образца Яндекса:
https://doc-static.yandex.net/src/support/webmaster/ru/new-doctors.xml,
скачан и сверен вручную 2026-09-04 - тег <shop>/<company>/<clinics>/<doctors>/
<services>/<offers>, НЕ то же самое, что колонки xlsx-шаблона).

Единственный источник правды по данным - xlsx (там уже проверенные цены из
CRM, флаги "Вызов на дом"/"Базовая услуга" и т.п.) - этот скрипт только
меняет формат представления, чтобы можно было проверить фид в
"XML-валидаторе" Вебмастера (категория "Врачи") ДО отправки на модерацию,
а не только после неё.

Запуск: python3 scripts/xlsx-to-yandex-doctors-yml.py
Результат: yandex-doctors-feed/biorise-doctors-feed.yml
"""

import os
from datetime import datetime
from xml.etree.ElementTree import Element, SubElement, ElementTree, indent
from xml.sax.saxutils import escape

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(ROOT, 'yandex-doctors-feed', 'biorise-doctors-feed.xlsx')
YML_PATH = os.path.join(ROOT, 'yandex-doctors-feed', 'biorise-doctors-feed.yml')


def to_bool_str(v):
    return 'true' if str(v).strip() == 'ИСТИНА' else 'false'


def to_international_phone(v):
    if v is None:
        return None
    digits = ''.join(ch for ch in str(v) if ch.isdigit())
    if not digits:
        return None
    if len(digits) == 11 and digits[0] == '8':
        digits = '7' + digits[1:]
    return f'+{digits}'


def sub(parent, tag, text):
    if text is None or text == '':
        return None
    el = SubElement(parent, tag)
    el.text = str(text)
    return el


def build():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    org_ws = wb['Данные организации']
    company_name = org_ws['B1'].value
    platform_name = org_ws['B2'].value
    site_url = org_ws['B3'].value
    company_email = org_ws['B4'].value
    company_logo = org_ws['B5'].value

    shop = Element('shop', {'version': '2.0', 'date': datetime.now().strftime('%Y-%m-%d %H:%M')})
    sub(shop, 'name', platform_name)
    sub(shop, 'company', company_name)
    sub(shop, 'url', site_url)
    sub(shop, 'picture', company_logo)
    sub(shop, 'email', company_email)

    # --- Врачи: собираем уникальных докторов (по col2 "Идентификатор врача") ---
    doctors_ws = wb['Врачи']
    doctors_seen = {}
    rows = []
    for row in doctors_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rows.append(row)
        doctor_id = row[1]
        if doctor_id not in doctors_seen:
            doctors_seen[doctor_id] = {
                'name': row[0],
                'description': row[9],
                'picture': row[11],
                'experience_years': row[12],
            }

    doctors_el = SubElement(shop, 'doctors')
    for doctor_id, d in doctors_seen.items():
        doctor_el = SubElement(doctors_el, 'doctor', {'id': str(doctor_id)})
        sub(doctor_el, 'internal_id', doctor_id)
        sub(doctor_el, 'name', d['name'])
        sub(doctor_el, 'description', d['description'])
        sub(doctor_el, 'picture', d['picture'])
        if d['experience_years']:
            sub(doctor_el, 'experience_years', int(d['experience_years']))

    # --- Клиники ---
    clinics_ws = wb['Данные клиник']
    clinics_el = SubElement(shop, 'clinics')
    for row in clinics_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name, clinic_id, url, email, city, address, phone, company_id, internal_id, picture = row
        clinic_el = SubElement(clinics_el, 'clinic', {'id': str(clinic_id)})
        sub(clinic_el, 'url', url)
        sub(clinic_el, 'picture', picture)
        sub(clinic_el, 'name', name)
        sub(clinic_el, 'city', city)
        sub(clinic_el, 'address', address)
        sub(clinic_el, 'email', email)
        sub(clinic_el, 'phone', to_international_phone(phone))
        sub(clinic_el, 'internal_id', internal_id)
        sub(clinic_el, 'company_id', company_id)

    # --- Услуги: id = внутренний идентификатор (уже уникален и латиницей) ---
    services_ws = wb['Услуги']
    services_el = SubElement(shop, 'services')
    service_id_by_name = {}
    for row in services_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name, _unused, description, internal_id = row
        service_id_by_name[name] = internal_id
        service_el = SubElement(services_el, 'service', {'id': str(internal_id)})
        sub(service_el, 'name', name)
        sub(service_el, 'description', description)
        sub(service_el, 'internal_id', internal_id)

    # --- Офферы: один <offer> на строку листа "Врачи" ---
    offers_el = SubElement(shop, 'offers')
    for i, row in enumerate(rows, start=1):
        (
            doc_name, doctor_id, doctor_card_url, price, discount, discount_terms,
            service_name, is_base_service_col, booking_url, description, speciality,
            picture, experience_years, clinic_id, reviews_count, degree, rank,
            category, other_titles, adult, children, phone_for_booking, oms,
            house_call, telemed, appointment, online_schedule, internal_doctor_id,
        ) = row

        offer_el = SubElement(offers_el, 'offer', {'id': f'offer_{i}'})
        sub(offer_el, 'url', booking_url)
        sub(offer_el, 'oms', to_bool_str(oms))
        sub(offer_el, 'online_schedule', to_bool_str(online_schedule))
        sub(offer_el, 'appointment', to_bool_str(appointment))

        if price:
            price_el = SubElement(offer_el, 'price')
            sub(price_el, 'base_price', int(price))
            sub(price_el, 'currency', 'RUB')

        service_id = service_id_by_name.get(service_name)
        if service_id:
            SubElement(offer_el, 'service', {'id': str(service_id)})

        clinic_ref = SubElement(offer_el, 'clinic', {'id': str(clinic_id)})
        doctor_ref = SubElement(clinic_ref, 'doctor', {'id': str(doctor_id)})
        sub(doctor_ref, 'speciality', speciality)
        sub(doctor_ref, 'children_appointment', to_bool_str(children))
        sub(doctor_ref, 'adult_appointment', to_bool_str(adult))
        sub(doctor_ref, 'house_call', to_bool_str(house_call))
        sub(doctor_ref, 'telemed', to_bool_str(telemed))
        sub(doctor_ref, 'is_base_service', 'true' if is_base_service_col == 'Да' else 'false')

    tree = ElementTree(shop)
    indent(tree, space='  ')
    tree.write(YML_PATH, encoding='utf-8', xml_declaration=True)
    print(f'Готово: {YML_PATH}')
    print(f'Докторов: {len(doctors_seen)}, клиник: {clinics_ws.max_row - 1}, услуг: {len(service_id_by_name)}, офферов: {len(rows)}')


if __name__ == '__main__':
    build()
