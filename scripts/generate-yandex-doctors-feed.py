#!/usr/bin/env python3
"""
Собирает фид "Врачи" для Яндекс.Вебмастера (Услуги и предложения -> Врачи)
из официального шаблона (yandex-doctors-feed/template.xlsx), данных сайта
и цен из Klientiks CRM (klientiks-pipeline/biorise-crm-export.json -
обнови её через `node klientiks-pipeline/fetch-services.mjs &&
python3 klientiks-pipeline/export.py`, если цены могли измениться).

Запуск: python3 scripts/generate-yandex-doctors-feed.py
Результат: yandex-doctors-feed/biorise-doctors-feed.xlsx
"""

import json
import os
import re
import shutil
import zipfile

import openpyxl

XML_DECLARATION = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'


def restore_xml_declarations(path):
    """openpyxl's writer omits the leading <?xml ...?> declaration from every
    internal part when it re-saves an .xlsx (verified by diffing our output
    against the untouched Yandex template byte-for-byte - every xl/*.xml part
    in the template starts with the declaration, ours starts with the bare
    tag). That's technically still well-formed XML, but Yandex's own feed
    validator is strict about it and rejects the file with "Лишние символы в
    начале фида, строка 1 позиция 1" - so re-inject the declaration into every
    XML part that's missing one before shipping the file."""
    tmp_path = path + '.tmp'
    with zipfile.ZipFile(path, 'r') as src, zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename.endswith('.xml') or item.filename.endswith('.rels'):
                if not data.startswith(b'<?xml'):
                    data = XML_DECLARATION + data
            dst.writestr(item, data)
    os.replace(tmp_path, path)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(ROOT, 'yandex-doctors-feed', 'template.xlsx')
OUTPUT_PATH = os.path.join(ROOT, 'yandex-doctors-feed', 'biorise-doctors-feed.xlsx')
SITE_URL = 'https://biorise-clinic.ru'
PHONE = '8 996 749 97 47'

with open(os.path.join(ROOT, 'klientiks-pipeline', 'biorise-crm-export.json'), encoding='utf-8') as f:
    CRM = json.load(f)

# --- Организация ---
ORGANIZATION = {
    'Название компании': 'ООО «МК Клиники Будущего»',
    'Название площадки': 'BIORISE',
    'Ссылка на сайт компании': SITE_URL,
    'Электронная почта компании': '',
    'Ссылка на логотип компании': f'{SITE_URL}/logo-cube.png',
}

# --- Клиники: только реально действующие адреса ---
# 'Название клиники' должно совпадать с названием карточки организации в Яндекс Бизнесе
# (проверено 2026-09-02: обе карточки называются просто "Биорайз", без адреса/бренда
# латиницей — фид с "BIORISE на Дыбенко" отклонён модерацией с формулировкой "удалите лишнее").
CLINICS = [
    {
        'id': 'clinic_dybenko',
        'Название клиники': 'Биорайз',
        'Город клиники': 'Самара',
        'Адрес клиники': 'ул. Дыбенко 27Б',
    },
    {
        'id': 'clinic_stara_zagora',
        'Название клиники': 'Биорайз',
        'Город клиники': 'Самара',
        'Адрес клиники': 'ул. Стара Загора 48',
    },
]
DYBENKO = 'clinic_dybenko'
STARA_ZAGORA = 'clinic_stara_zagora'

# --- Врачи (продублировано из lib/authors.ts) ---
# clinics: на каких адресах реально принимает этот врач.
DOCTORS = {
    'potemkina': {
        'name': 'Потемкина Ольга Владимировна',
        'specialty': 'терапевт',
        'experience_years': 19,
        'bio': 'Ведёт приём как врач-терапевт со стажем 19 лет.',
        'avatar': '/optimized/doctors/potemkina.webp',
        'clinics': [DYBENKO, STARA_ZAGORA],
        # Цена из Klientiks CRM ("Прием (осмотр, консультация) врача-терапевта первичный").
        'base_service': ('Первичный приём терапевта', 2200),
    },
    'boboeva': {
        'name': 'Бобоева Наталья',
        'specialty': 'эндокринолог',
        'experience_years': None,
        'bio': 'Врач-эндокринолог. Работает с хроническими эндокринными заболеваниями.',
        'avatar': '/optimized/doctors/boboeva.webp',
        'clinics': [DYBENKO, STARA_ZAGORA],
        'base_service': ('Первичный приём эндокринолога', 2800),
    },
    'tregubova': {
        'name': 'Трегубова Лиана Игоревна',
        'specialty': 'диетолог',
        'experience_years': None,
        'bio': 'Врач-диетолог, специализируется на интегративной превентивной и антивозрастной медицине.',
        'avatar': '/optimized/doctors/tregubova.webp',
        'clinics': [DYBENKO, STARA_ZAGORA],
        'base_service': ('Первичный приём диетолога', 2200),
    },
    'malofeeva': {
        'name': 'Малофеева Кристина Владимировна',
        'specialty': 'косметолог',
        'experience_years': None,
        'bio': 'Косметолог-эстетист, мастер лазерной эпиляции.',
        'avatar': '/optimized/doctors/malofeeva.webp',
        'clinics': [DYBENKO],  # на Стара Загоре косметолога нет
        # В CRM нет отдельной услуги "первичный приём косметолога" - только конкретные
        # процедуры. Это тот самый документированный случай специальности, где базовая
        # услуга неприменима (см. yandex.ru/support/webmaster/.../doctors), поэтому
        # base_service не задаём - у неё все предложения останутся доп. услугами.
        'base_service': None,
    },
    'zakharov': {
        'name': 'Захаров Александр Владимирович',
        'specialty': 'массажист',
        'experience_years': None,
        'bio': 'Косметик-эстетист по уходу за телом, массажист.',
        'avatar': '/optimized/doctors/zakharov.webp',
        'clinics': [DYBENKO],  # массаж только на Дыбенко
        'base_service': None,  # то же самое: у массажа нет обобщённой "консультации"
    },
}

# --- Капельницы: слаг сайта -> "короткое" имя для сопоставления с CRM ---
# (скопировано из scripts/generate-kapelnicy.js: nameToPriceName, ключ - title сайта)
TITLE_TO_PRICE_NAME = {
    'Капельница Иммуносуппорт': 'Иммуносуппорт',
    'Капельница с витаминами': 'Витаминная',
    'Капельница Коктейль Майерса': 'Мультивитаминная',
    'Капельница при ОРВИ': 'Антивирус',
    'Капельница после ковида': 'Постковид',
    'Капельница «Детокс»': 'Детокс стандарт',
    'Капельница для печени': 'Детоксикация',
    'Капельница при отравлении': 'Детоксикация',
    'Капельница с Гептралом': 'Детоксикация',
    'Капельница от стресса и нервов': 'Антистресс',
    'Капельница для мозга': 'Брейнсторм',
    'Капельница для сосудов': 'Здоровые сосуды',
    'Капельница для сердца': 'Здоровые сосуды',
    'Капельницы для спортсменов': 'Спорт силовая',
    'Спорт кардио': 'Спорт кардио',
    'Протеин буст': 'Протеин буст',
    'Капельница для похудения': 'Снижение веса',
    'Капельница при диабете': 'Сахар в норме',
    'Капельница «Золушка»': 'Золушка (Красота и молодость)',
    'Капельница с глутатионом': 'Антиэйдж премиум',
    'Капельница с железом': 'Железо стандарт',
    'Капельница при беременности': 'Маме можно',
    'Капельница при аллергии': 'Антигистаминная',
    'Половая система': 'Половая система',
    'Мужское здоровье': 'Мужское здоровье',
    'Капельница с глюкозой': 'После вечеринки',
    'Капельница Лаеннек': 'Лаеннек',
    'Капельница Феринжект': 'Железо 2.0',
    'Капельница Антиаммиак': 'Антиаммиак',
    'Капельница Антиклимакс': 'Анти-климакс',
    'Капельница Антимигрень': 'Антимигрень',
    'Капельница Айронмен': 'Айронмен',
    'Капельница Бархатная кожа': 'Бархатная кожа',
    'Капельница бронхо-легочная противовоспалительная': 'Бронхолегочная',
    'Капельница Детоксикация': 'Детоксикация',
    'Капельница Джетлаг': 'Джетлаг',
    'Капельница Гинекологическая противовоспалительная': 'Гинекологическая противоспалительная',
    'Капельница Гипертонический криз': 'Гипертонический криз',
    'Капельница Густые волосы': 'Густые волосы',
    'Капельница Нет холестерину': 'Нет холестерину',
    'Капельница Подготовка к беременности': 'Подготовка к беременности',
    'Капельница Предоперационная': 'Предоперационная',
    'Капельница Стоматологическая противовоспалительная': 'Стоматологическая противоспалительная',
    'Капельница суставная противовоспалительная': 'Суставная противоспалительная',
    'Капельница Восстановление обоняния': 'Восстановление обоняния',
    'Капельница ЖКТ 1': 'ЖКТ 1',
    'Капельница ЖКТ 2': 'ЖКТ 2',
    'Терзапатид 10 мг': 'Терзапатид 10 мг',
    'Капельница Энергия+': 'Энергия +',  # в самом сайте это сейчас БАГ (нет пробела) - цену берём из CRM напрямую
}

KAPELNICY_DOCTOR_OVERRIDES = {
    'sahar-v-norme': 'boboeva',
    'antiklimaks': 'boboeva',
    'snizhenie-vesa': 'tregubova',
    'terzapatid': 'tregubova',
    'detoks-standart': 'tregubova',
    'detoksikatsiya-pechen': 'tregubova',
    'detoksikatsiya-otravlenie': 'tregubova',
    'detoksikatsiya-geptral': 'tregubova',
    'detoksikatsiya': 'tregubova',
    'antiammiak': 'tregubova',
    'krasota-i-omolozhenie': 'tregubova',
    'antieydzh-premium': 'tregubova',
    'barhatnaya-kozha': 'malofeeva',
    'gustye-volosy': 'malofeeva',
}

# --- Чек-апы: слаг -> подстрока для поиска в CRM (продублировано из lib/checkups.ts) ---
CHECKUPS = [
    ('zhenskoe-zdorove-bazovyy', 'Женское здоровье: базовый чек-ап', 'женское здоровье базовый'),
    ('muzhskoe-zdorove-bazovyy', 'Мужское здоровье: базовый чек-ап', 'мужское здоровье базовый'),
    ('detskiy-bazovyy', 'Детский чек-ап: базовый', 'детский'),
    ('diagnostika-defitsita-zheleza', 'Чек-ап Диагностика дефицита железа', 'диагностика дефицита железа'),
    ('zhenskoe-zdorove-optimal', 'Женское здоровье Optimal', 'женское здоровье optimal'),
    ('muzhskoe-zdorove-optimal', 'Мужское здоровье Optimal', 'мужское здоровье optimal'),
    ('zhenskoe-zdorove-premium', 'Женское здоровье Premium', 'женское здоровье  premium'),
    ('muzhskoe-zdorove-premium', 'Мужское здоровье Premium', 'мужское здоровье premium'),
    ('detskiy-podgotovka-k-shkole', 'Детский чек-ап: подготовка к школе и саду', None),  # нет в CRM отдельной строкой
    ('zabota-o-roditelyakh', 'Чек-ап Забота о родителях', 'забота о родителях'),
]
CHECKUPS_FALLBACK_PRICE = {'detskiy-podgotovka-k-shkole': 1490}
CHECKUPS_KIDS_SLUGS = {'detskiy-bazovyy', 'detskiy-podgotovka-k-shkole'}

# --- Отдельные анализы со своей страницей (подтверждено против CRM 2026-08-31) ---
STANDALONE_ANALYSES = [
    ('analiz-krovi', 'Общий анализ крови', 270),
    ('biohimicheskiy-analiz-krovi', 'Биохимический анализ крови', 1550),
    ('analiz-mochi', 'Общий анализ мочи', 280),
    ('ferritin', 'Ферритин', 520),
    ('vitamin-d', 'Витамин D', 2065),
    ('t-spot', 'T-SPOT', 6500),
]


def norm(s):
    s = (s or '').lower().strip()
    s = s.replace('«', '').replace('»', '').replace('"', '')
    s = re.sub(r'\s+', ' ', s)
    return s


def norm_kap_name(s):
    s = norm(s)
    s = re.sub(r'^капельниц[а-я]*\s*', '', s)
    return s.strip()


def load_kapelnicy():
    with open(os.path.join(ROOT, 'data', 'kapelnicy.generated.json'), encoding='utf-8') as f:
        return json.load(f)


# Опечатки в самих названиях в CRM (не в нашем справочнике) - ручные исключения.
CRM_NAME_TYPO_OVERRIDES = {
    'нет холестерину': 'капельница капельница нет холестирину',
    'стоматологическая противоспалительная': 'капельница стомотологическая противоспалительная',
    'терзапатид 10 мг': 'терзапатид 10мг',
}


def crm_price_by_exact_name(group, exact_name):
    target = norm_kap_name(exact_name)
    override = CRM_NAME_TYPO_OVERRIDES.get(target)
    for s in CRM['services']:
        name_norm = norm(s['Название'])
        if override and name_norm == override:
            return s['Цена, ₽']
        if s['Группа'] == group and norm_kap_name(s['Название']) == target:
            return s['Цена, ₽']
    return None


def crm_checkup_price(needle):
    if needle is None:
        return None
    needle = norm(needle)
    for s in CRM['services']:
        if s['Группа'] == 'Чек ап' and needle in norm(s['Название']):
            return s['Цена, ₽']
    return None


def crm_group_items(groups):
    return [s for s in CRM['services'] if s['Группа'] in groups and not s['В архиве']]


def build():
    kapelnicy = load_kapelnicy()

    shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)
    wb = openpyxl.load_workbook(OUTPUT_PATH)

    # --- Данные организации ---
    # Этот лист устроен вертикально: колонка A - подпись поля (не трогать),
    # колонка B - значение, ряд на ряд (5 полей = 5 строк), а не горизонтальная
    # таблица "заголовок + одна строка данных". Раньше сюда дописывалась вторая
    # строка A2:E2 - реальные данные никогда не подменяли пример в B1:B5, поэтому
    # модерация Яндекса читала оттуда "ООО Яндекс" и ссылку-заглушку на логотип.
    ws = wb['Данные организации']
    ws['B1'] = ORGANIZATION['Название компании']
    ws['B2'] = ORGANIZATION['Название площадки']
    ws['B3'] = ORGANIZATION['Ссылка на сайт компании']
    ws['B4'] = ORGANIZATION['Электронная почта компании']
    ws['B5'] = ORGANIZATION['Ссылка на логотип компании']

    # --- Данные клиник ---
    ws = wb['Данные клиник']
    row = 2
    for clinic in CLINICS:
        ws.cell(row=row, column=1, value=clinic['Название клиники'])
        ws.cell(row=row, column=2, value=clinic['id'])
        ws.cell(row=row, column=3, value=SITE_URL)
        ws.cell(row=row, column=4, value='')
        ws.cell(row=row, column=5, value=clinic['Город клиники'])
        ws.cell(row=row, column=6, value=clinic['Адрес клиники'])
        ws.cell(row=row, column=7, value=PHONE)
        ws.cell(row=row, column=8, value='')
        ws.cell(row=row, column=9, value=clinic['id'])
        ws.cell(row=row, column=10, value=f'{SITE_URL}/logo-cube.png')
        row += 1

    services_ws = wb['Услуги']
    doctors_ws = wb['Врачи']
    srow = 2
    drow = 2
    unmatched_report = []

    def add_service(internal_id, name, description):
        nonlocal srow
        services_ws.cell(row=srow, column=1, value=name)
        services_ws.cell(row=srow, column=2, value='')
        services_ws.cell(row=srow, column=3, value=description)
        services_ws.cell(row=srow, column=4, value=internal_id)
        srow += 1

    def add_offer(doctor_id, clinic_id, service_name, price, booking_url, is_kids=False, home_visit=False, is_base=False):
        nonlocal drow
        doc = DOCTORS[doctor_id]
        values = [
            doc['name'], doctor_id, f'{SITE_URL}/vrachi/{doctor_id}/',
            price, '', '', service_name, 'Да' if is_base else 'Нет', booking_url,
            doc['bio'], doc['specialty'], f'{SITE_URL}{doc["avatar"]}',
            doc['experience_years'] if doc['experience_years'] else '',
            clinic_id, '', '', '', '', '',
            'ИСТИНА', 'ИСТИНА' if is_kids else 'ЛОЖЬ',
            PHONE, 'ЛОЖЬ', 'ИСТИНА' if home_visit else 'ЛОЖЬ', 'ЛОЖЬ', 'ИСТИНА', 'ЛОЖЬ',
            doctor_id,
        ]
        for col, value in enumerate(values, start=1):
            doctors_ws.cell(row=drow, column=col, value=value)
        drow += 1

    def add_offer_all_clinics(doctor_id, service_name, price, booking_url, is_kids=False, home_visit=False, is_base=False):
        for clinic_id in DOCTORS[doctor_id]['clinics']:
            add_offer(doctor_id, clinic_id, service_name, price, booking_url, is_kids, home_visit, is_base)

    # --- Базовая услуга: "для каждого врача/клиники/специальности должно быть ровно
    # одно предложение с базовой услугой (обычно первичный приём)" (требование Яндекса).
    # Раньше 'Да' стояло у КАЖДОЙ строки (у каждой капельницы/анализа) - это нарушение.
    for doctor_id, doc in DOCTORS.items():
        if not doc.get('base_service'):
            continue
        name, price = doc['base_service']
        add_service(f'base_{doctor_id}', name, name)
        add_offer_all_clinics(doctor_id, name, price, f'{SITE_URL}/vrachi/{doctor_id}/', is_base=True)

    # --- Капельницы ---
    for slug, item in kapelnicy.items():
        title = item.get('title')
        if not title:
            continue
        price_name = TITLE_TO_PRICE_NAME.get(title)
        price = crm_price_by_exact_name('Капельницы', price_name) if price_name else None
        if price is None:
            site_price = ''.join(ch for ch in (item.get('price') or '') if ch.isdigit())
            price = int(site_price) if site_price else None
            if price is not None:
                unmatched_report.append((slug, title, 'CRM: не найдено, использована цена сайта', price))
        if price is None:
            unmatched_report.append((slug, title, 'CRM: не найдено, цены на сайте тоже нет - ПРОПУЩЕНО', None))
            continue
        doctor_id = KAPELNICY_DOCTOR_OVERRIDES.get(slug, 'potemkina')
        add_service(f'kap_{slug}', title, item.get('description', ''))
        # Весь каталог капельниц доступен с выездом на дом (см. /kapelnicy/na-domu/,
        # 2500 руб. за выезд + стоимость капельницы, решение о безопасности - за врачом).
        add_offer_all_clinics(doctor_id, title, price, f'{SITE_URL}/kapelnicy/{slug}/', home_visit=True)

    # --- Чек-апы ---
    for slug, title, needle in CHECKUPS:
        price = crm_checkup_price(needle) or CHECKUPS_FALLBACK_PRICE.get(slug)
        if price is None:
            unmatched_report.append((slug, title, 'CRM Чек ап: не найдено - ПРОПУЩЕНО', None))
            continue
        add_service(f'chekap_{slug}', title, f'Программа обследования: {title}')
        add_offer_all_clinics('potemkina', title, price, f'{SITE_URL}/chek-apy/{slug}/', is_kids=slug in CHECKUPS_KIDS_SLUGS)

    # --- Отдельные анализы ---
    for slug, title, price in STANDALONE_ANALYSES:
        add_service(f'analiz_{slug}', title, f'Лабораторный анализ: {title}')
        add_offer_all_clinics('potemkina', title, price, f'{SITE_URL}/analizy/{slug}/')

    # --- Массаж (Захаров), только Дыбенко ---
    for s in crm_group_items(['Медицинский массаж', 'аппаратный массаж']):
        name = s['Название']
        price = s['Цена, ₽']
        add_service(f'massage_{s["ID"]}', name, name)
        add_offer('zakharov', DYBENKO, name, price, f'{SITE_URL}/ruchnoy-massazh/')

    # --- Косметология + лазерная эпиляция (Малофеева), только Дыбенко ---
    for s in crm_group_items(['косметология', 'Лазерная эпиляция']):
        name = s['Название']
        price = s['Цена, ₽']
        add_service(f'cosmetology_{s["ID"]}', name, name)
        add_offer('malofeeva', DYBENKO, name, price, f'{SITE_URL}/lazernaya-epilyatsiya/')

    # Опциональные листы содержат в шаблоне готовую строку-пример (фейковый врач
    # "Orlov-1" со ссылками на doctors.sample.s3.yandex.net) - мы их не используем,
    # но если оставить пример как есть, Яндекс попытается провалидировать его как
    # настоящую запись и отклонит по недопустимому домену картинки.
    for sheet_name in ('Отзывы о враче', 'Доп информация врача'):
        opt_ws = wb[sheet_name]
        for row in opt_ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.value = None

    wb.save(OUTPUT_PATH)
    restore_xml_declarations(OUTPUT_PATH)
    print(f'Готово: {OUTPUT_PATH}')
    print(f'Услуг: {srow - 2}, строк-предложений: {drow - 2}')
    print()
    print('=== Несовпадения/пропуски при сопоставлении с CRM ===')
    for slug, title, note, price in unmatched_report:
        print(f'{slug} | "{title}" | {note} | цена={price}')


if __name__ == '__main__':
    build()
